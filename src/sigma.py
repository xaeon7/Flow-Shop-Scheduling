from os import read
import sys
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import QApplication, QMainWindow, QStackedWidget,QHeaderView,QFileDialog,QTableWidgetItem,QTableWidget, QWidget,QVBoxLayout
from openpyxl import load_workbook
import excelScripts

from PyQt5 import QtCore
from PyQt5 import QtGui

import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('Qt5Agg')
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg, NavigationToolbar2QT 
import matplotlib.figure as Figure
import numpy as np

import seaborn as sns
import pandas as pd
import sip

from lib.CDS import FlowShopSchedule

colors = ['#FEB800', '#33D6A0', '#6F52ED', '#FE4C61', '#06A6FF', '#FF7A01', '#FF66B3', '#4DD1BA', '#FA656A', '#477CFF']



class MatplotlibCanvas(FigureCanvasQTAgg):
    def __init__(self, parent = None, width=5, height = 5, dpi = 90):
        fig, self.gnt = plt.subplots(figsize = (width, height), dpi= dpi)
        super(MatplotlibCanvas, self).__init__(fig)     
        fig.patch.set_facecolor('#F8F7FC')
        fig.tight_layout()

class StartScreen(QMainWindow):
    def __init__(self):
        super(StartScreen, self).__init__()
        loadUi("screens/start.ui",self)
        self.start.clicked.connect(self.gotojobsInput)
        self.start.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.algorithm.setEnabled(False)
        self.optimize.setEnabled(False)
        self.preparation.clicked.connect(self.updateAlgo)
        self.delai.clicked.connect(self.updateOptimize)

        
    def updateAlgo(self):
        self.algorithm.setEnabled(self.preparation.isChecked())
        
    def updateOptimize(self):
        self.optimize.setEnabled(self.delai.isChecked())

    def gotojobsInput(self):
        widget.nb_jobs=self.nb_jobs.value()
        widget.nb_machines=self.nb_machines.value()
        widget.delay=self.delai.isChecked()
        widget.prep=self.preparation.isChecked()
        widget.algo = self.algorithm.currentIndex()
        widget.optimize = self.optimize.currentIndex()
        inputJobs=JobsMatrixInput()
        widget.addWidget(inputJobs)
        widget.setCurrentIndex(widget.currentIndex()+1)


class JobsMatrixInput(QMainWindow):
    def __init__(self):
        super(JobsMatrixInput, self).__init__()
        loadUi("screens/jobsMatrix.ui",self)
        self.createTable(columns=widget.nb_jobs,
                        rows=widget.nb_machines)
        

        self.importButton.clicked.connect(lambda :self.importData(2,2))
        self.importButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.continueButton.clicked.connect(self.goNext)
        self.continueButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.backButton.clicked.connect(self.goBack)
        self.backButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.logo.clicked.connect(self.goHome)
        self.continueButton.setStyleSheet("border-radius: 10px;\n"
"font: 14pt \"Montserrat SemiBold\";\n"
"color:white;\n"
"  background-color: rgb(181, 181, 204);")
        self.continueButton.setEnabled(False)
        self.matrixTable.itemChanged.connect(self.updateButton)



    def createTable(self,columns,rows):
        widget.nb_jobs=columns
        widget.nb_machines=rows
        widget.jobsMatrix = []
        self.matrixTable.setRowCount(rows)
        self.matrixTable.setColumnCount(columns)
        self.matrixTable.clear()
        self.matrixTable.setHorizontalHeaderLabels(["Job "+str(i+1) for i in range(columns)])
        self.matrixTable.setVerticalHeaderLabels(["Machine "+str(i+1) for i in range(rows)])
        self.matrixTable.horizontalHeader().setStretchLastSection(True)
        self.matrixTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    # def importData(self,initCol,initRow):
    #     options = QFileDialog.Options()
    #     options |= QFileDialog.DontUseNativeDialog
    #     fileName, _ = QFileDialog.getOpenFileName(self,"Import Jobs Data","", "Excel Files (*.xlsx)", options=options)
    #     if not(fileName):
    #         return 0
    #     wb2 = load_workbook(fileName)
    #     sheet=wb2['Feuil1']
    #     matrix=excelScripts.readMatrix(sheet,initCol,initRow)
    #     self.createTable(rows=len(matrix),columns=len(matrix[0]))
    #     self.fillTable(matrix)
        
    def importData(self,initCol,initRow):
        self.errorType = 0
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"Import Jobs Data", "","Excel Files (*.xlsx)", options=options)
        if not(fileName):
            return 0
        wb2 = load_workbook(fileName)
        sheet=wb2.worksheets[0]
        matrix=excelScripts.readMatrix(sheet,initCol,initRow)
        self.createTable(rows=len(matrix),columns=len(matrix[0]))
        self.fillTable(matrix)
        
            
    def updateButton(self):
        check= all([all([self.matrixTable.item(row, col) for col in range(widget.nb_jobs)]) for row in range(widget.nb_machines)])
        if check:
            self.continueButton.setStyleSheet("border-radius: 10px;\n"
"font: 14pt \"Montserrat SemiBold\";\n"
"color:white;\n"
"  background-color: #6F51EC;")
            self.continueButton.setEnabled(True)
        else:
            self.continueButton.setStyleSheet("border-radius: 10px;\n"
"font: 14pt \"Montserrat SemiBold\";\n"
"color:white;\n"
"  background-color: rgb(181, 181, 204);")
            self.continueButton.setEnabled(False)
            self.errorType = 0
    
    def fillTable(self,matrix):
        try:
            for i in range(widget.nb_machines):
                for j in range(widget.nb_jobs):
                    self.matrixTable.setItem(i,j, QTableWidgetItem(str(matrix[i][j])))
        except:
            self.errorType = 1

    def readTable(self):
        jobsMatrix=[[int(self.matrixTable.item(row, col).text()) for col in range(widget.nb_jobs)] for row in range(widget.nb_machines)]
        widget.jobsMatrix = jobsMatrix
    
    def goBack(self):
        widget.setCurrentIndex(widget.currentIndex()-1)
        widget.removeWidget(self)
        self.errorType = 0
        
    def goHome(self):
        widget.setCurrentIndex(0)
        self.errorType = 0
        
        for i in range(widget.count(), 0, -1):
            widget.removeWidget(widget.widget(i))

    def goNext(self):
        try:
            self.readTable()
        except:
            self.errorType = 1 
        
        if widget.delay:
            self.gotoDelayInput()
            
        elif widget.prep:
            self.gotoPrepInput()
                      
        elif self.errorType == 0:
            try:
                if widget.jobsMatrix :
                    self.gotoGantt()
                else:
                    self.errorType = 1
                    self.goNext()
            except:
                self.errorType = 1
        
        else:
            self.error.setText('The data is invalid.')

    def gotoDelayInput(self):
        inputDelay=DelayInput()
        widget.addWidget(inputDelay)
        widget.setCurrentIndex(widget.currentIndex()+1)
    
    def gotoPrepInput(self):
        inputPrep=PrepInput()
        widget.addWidget(inputPrep)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def gotoGantt(self):
        gantt=DisplayGantt()
        widget.addWidget(gantt)
        widget.setCurrentIndex(widget.currentIndex()+1)

class DelayInput(QMainWindow):
    def __init__(self):
        super(DelayInput, self).__init__()
        loadUi("screens/delay.ui",self)
        self.dataLabel.setText("Delay of jobs")
        self.loadTable()
        

        self.importButton.clicked.connect(lambda :self.importDelai(2,2))
        self.importButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.continueButton.clicked.connect(self.goNext)
        self.continueButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.backButton.clicked.connect(self.goBack)
        self.importButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.logo.clicked.connect(self.goHome)
        self.continueButton.setStyleSheet("border-radius: 10px;\n"
"font: 14pt \"Montserrat SemiBold\";\n"
"color:white;\n"
"  background-color: rgb(181, 181, 204);")
        self.continueButton.setEnabled(False)
        self.matrixTable.itemChanged.connect(self.updateButton)
        
        self.errorType = 0
            
        
    def updateButton(self):
        check= all([self.matrixTable.item(0, col) for col in range(widget.nb_jobs)])
        if check:
            self.continueButton.setStyleSheet("border-radius: 10px;\n"
"font: 14pt \"Montserrat SemiBold\";\n"
"color:white;\n"
"  background-color: #6F51EC;")
            self.continueButton.setEnabled(True)
        else:
            self.continueButton.setStyleSheet("border-radius: 10px;\n"
"font: 14pt \"Montserrat SemiBold\";\n"
"color:white;\n"
"  background-color: rgb(181, 181, 204);")
            self.continueButton.setEnabled(False)  
              


    def loadTable(self):
        self.matrixTable.setRowCount(1)
        self.matrixTable.setColumnCount(widget.nb_jobs)
        self.matrixTable.clear()
        self.matrixTable.setHorizontalHeaderLabels(["Job "+str(i+1) for i in range(widget.nb_jobs)])
        self.matrixTable.setVerticalHeaderLabels(["Delay"])
        self.matrixTable.horizontalHeader().setStretchLastSection(True)
        self.matrixTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    # def importDelai(self,initCol,initRow):
    #     options = QFileDialog.Options()
    #     options |= QFileDialog.DontUseNativeDialog
    #     fileName, _ = QFileDialog.getOpenFileName(self,"Import Delay Data", "","Excel Files (*.xlsx)", options=options)
    #     if not(fileName):
    #         return 0
    #     wb2 = load_workbook(fileName)
    #     sheet=wb2['Feuil1']
    #     seq=excelScripts.readSeq(sheet,initCol,initRow)
    #     self.fillTable(seq)
    
    def importDelai(self,initCol,initRow):
        self.errorType = 0
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"Import Delay Data", "","Excel Files (*.xlsx)", options=options)
        if not(fileName):
            return 0
        wb2 = load_workbook(fileName)
        sheet=wb2.worksheets[0]
        seq=excelScripts.readSeq(sheet,initCol,initRow)
        self.fillTable(seq)
    
    def fillTable(self,seq):
        try:
            for j in range(widget.nb_jobs):
                self.matrixTable.setItem(0,j, QTableWidgetItem(str(seq[j])))
        except:
            self.errorType = 1
    
    def readTable(self):
        delaySeq=[int(self.matrixTable.item(0, col).text()) for col in range(widget.nb_jobs)]
        widget.delaySeq = delaySeq
    
    def goBack(self):
        widget.setCurrentIndex(widget.currentIndex()-1)
        widget.removeWidget(self)
        self.errorType = 0
        
    def goHome(self):
        widget.setCurrentIndex(0)
        self.errorType = 0
        
        for i in range(widget.count(), 0, -1):
            widget.removeWidget(widget.widget(i))
    
    def goNext(self):               
        try:
            self.readTable()
        except:
            self.errorType = 1 
        
        if widget.prep:
            try:
                self.gotoPrepInput()
            except:
                self.errorType = 1
                                   
        elif self.errorType == 0:
            try:
                if widget.jobsMatrix :
                    self.gotoGantt()
                else:
                    self.errorType = 1
                    self.goNext()
            except:
                self.errorType = 1
        else:
            self.error.setText('The data is invalid.')
                
    def gotoPrepInput(self):
        inputPrep=PrepInput()
        widget.addWidget(inputPrep)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def gotoGantt(self):
        gantt=DisplayGantt()
        widget.addWidget(gantt)
        widget.setCurrentIndex(widget.currentIndex()+1)

class PrepInput(QMainWindow):
    def __init__(self):
        super(PrepInput, self).__init__()
        loadUi("screens/inputPrep.ui",self)
        self.loadTab()
        self.importButton.clicked.connect(lambda:self.importData(2,2))
        self.importButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.continueButton.clicked.connect(self.goNext)
        self.continueButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.backButton.clicked.connect(self.goBack)
        self.backButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.logo.clicked.connect(self.goHome)
        self.continueButton.setStyleSheet("border-radius: 10px;\n"
"font: 14pt \"Montserrat SemiBold\";\n"
"color:white;\n"
"  background-color: rgb(181, 181, 204);")
        self.continueButton.setEnabled(False)
        self.errorType = 0
        

        for mtab in range(widget.nb_machines):
            tab=self.prepTabs.widget(mtab).children()[1]
            tab.itemChanged.connect(self.updateButton)        

    def updateButton(self):
        
        flag = []
        for mtab in range(widget.nb_machines):
            tab=self.prepTabs.widget(mtab).children()[1]
            flag.append(all([tab.item(row, col)for col in range(widget.nb_jobs) for row in range(widget.nb_jobs)]))
        check= all(flag)
        
        if check:
            self.continueButton.setStyleSheet("border-radius: 10px;\n"
"font: 14pt \"Montserrat SemiBold\";\n"
"color:white;\n"
"  background-color: #6F51EC;")
            self.continueButton.setEnabled(True)
        else:
            self.continueButton.setStyleSheet("border-radius: 10px;\n"
"font: 14pt \"Montserrat SemiBold\";\n"
"color:white;\n"
"  background-color: rgb(181, 181, 204);")
            self.continueButton.setEnabled(False)   
            
    def loadTab(self):
        # generate a Tab for each machine and load tables
        n, m = widget.nb_jobs, widget.nb_machines
        self.prepMatrix=[[[]*n]*n]*m
        for j in range(m):
            table=QTableWidget()
            page=QWidget()
            page.layout = QVBoxLayout()
            page.layout.addWidget(table)
            page.setLayout(page.layout)
            self.prepTabs.addTab(page,"Machine "+str(j+1))
            tab=self.prepTabs.widget(j).children()[1]
            tab.setRowCount(n)
            tab.setColumnCount(n)
            tab.clear()
            headers=["Job "+str(i+1) for i in range(n)]
            tab.setHorizontalHeaderLabels(headers)
            tab.setVerticalHeaderLabels(headers)
            tab.horizontalHeader().setStretchLastSection(True)
            tab.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    # def importData(self,initCol,initRow):
    #     options = QFileDialog.Options()
    #     options |= QFileDialog.DontUseNativeDialog
    #     fileName, _ = QFileDialog.getOpenFileName(self,"Import Preparation Data", "","Excel Files (*.xlsx)", options=options)
    #     if not(fileName):
    #         return 0
    #     wb2 = load_workbook(fileName)
    #     sheet=wb2['Feuil1']
    #     prepMatrix=excelScripts.readprepMatrix(sheet,initCol,initRow)
    #     widget.prepMatrix=prepMatrix
    #     self.fillTabs()
    
    def importData(self,initCol,initRow):
        self.errorType = 0
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"Import Preparation Data", "","Excel Files (*.xlsx)", options=options)
        if not(fileName):
            return 0
        wb2 = load_workbook(fileName)
        widget.prepMatrix=[]
        for sheet in wb2.worksheets:
            matrix=excelScripts.readMatrix(sheet,initCol,initRow)
            widget.prepMatrix.append(matrix)
        self.fillTabs()
    
    def fillTabs(self):
        try:
            n, m = widget.nb_jobs, widget.nb_machines
            for mtab in range(m):
                tab=self.prepTabs.widget(mtab).children()[1]
                print(tab)
                for jobi in range(n):
                    for jobj in range(n):
                        tab.setItem(jobi,jobj, QTableWidgetItem(str(widget.prepMatrix[mtab][jobi][jobj])))
        except:
            self.errorType = 1

    
    def readTabs(self):
        n, m = widget.nb_jobs, widget.nb_machines
        widget.prepMatrix=[]
        for mtab in range(m):
            tab=self.prepTabs.widget(mtab).children()[1]
            tabData = [[int(tab.item(row, col).text()) for col in range(n)] for row in range(n)]
            widget.prepMatrix.append(tabData)
    
    def goBack(self):
        widget.setCurrentIndex(widget.currentIndex()-1)
        widget.removeWidget(self)
        self.errorType = 0
        

    def goHome(self):
        widget.setCurrentIndex(0)
        self.errorType = 0
        
        for i in range(widget.count(), 0, -1):
            widget.removeWidget(widget.widget(i))

    def goNext(self):
        try:
            self.readTabs()
        except:
            self.errorType = 1
                                   
        if self.errorType == 0:
            try:
                if widget.jobsMatrix :
                    self.gotoGantt()
                else:
                    self.errorType = 1
                    self.goNext()
            except:
                self.errorType = 1
            
        else:
            self.error.setText('The data is invalid.')
               
    def gotoGantt(self):
        gantt=DisplayGantt()
        widget.addWidget(gantt)
        widget.setCurrentIndex(widget.currentIndex()+1)
    

class DisplayGantt(QMainWindow):
    def __init__(self):
        super(DisplayGantt, self).__init__()
        loadUi("screens/plots.ui",self)
        
        self.flowShopSchedule = FlowShopSchedule(widget.jobsMatrix)
        self.logo.clicked.connect(self.goHome)

        
        if not widget.delay and not widget.prep:
            self.flowShopSchedule.CDS()
        
        elif widget.delay and not widget.prep:
            self.flowShopSchedule.Delay(widget.delaySeq, widget.optimize)
            
        elif not widget.delay and widget.prep:
            self.flowShopSchedule.Preparation(widget.prepMatrix, widget.algo)
            
        elif widget.delay and widget.prep:
            self.flowShopSchedule.PreparationAndDelay(widget.prepMatrix, widget.delaySeq, widget.algo)
        
        self.gridValue = False  
        
###########????????????????
        if not len(self.flowShopSchedule.stopPreparationRate):
            self.tap.setEnabled(False)  
            self.tap.setStyleSheet("color: \"#ffffff\";\n"
        "background-color: rgb(181, 181, 204);\n"
        "border: 1px solid rgb(181, 181, 204) ;\n"
        "border-top-left-radius:\'5px\';\n"
        "border-bottom-left-radius:\'5px\';\n"
        "border-top-right-radius:\'5px\';\n"
        "border-bottom-right-radius:\'5px\';\n"
        "font: 11pt \"Montserrat SemiBold\";\n"
        "width: \'95px\';\n"
        "height: \'30px\';")
        else:
                self.tap.setEnabled(True) 
                self.tap.setStyleSheet("color: \"#ffffff\";\n"
        "background-color: rgb(111, 81, 237);\n"
        "border: 1px solid rgb(111, 81, 237) ;\n"
        "border-top-left-radius:\'5px\';\n"
        "border-bottom-left-radius:\'5px\';\n"
        "border-top-right-radius:\'5px\';\n"
        "border-bottom-right-radius:\'5px\';\n"
        "font: 11pt \"Montserrat SemiBold\";\n"
        "width: \'95px\';\n"
        "height: \'30px\';")
                
###########????????????????

########??? TOOLBAR ###########################################
        self.canv = MatplotlibCanvas(self)
        self.toolbar = NavigationToolbar2QT(self.canv, self.centralwidget)
        self.tarTB.addWidget(self.toolbar)
        self.tfrTB.addWidget(self.toolbar)
        self.tapTB.addWidget(self.toolbar)
        self.ganttTB.addWidget(self.toolbar)
########???###########################################


        
                
########??? Buttons ##################################       
        self.back.clicked.connect(self.goBack)   
        self.gantt.clicked.connect(self.displayGantt)     
        self.tfr.clicked.connect(self.displayTFR)     
        self.tap.clicked.connect(self.displayTAP)           
        self.tar.clicked.connect(self.displayTAR)
        self.checkBox.clicked.connect(self.displayGrid)
########???###########################################
        
        self.displayGantt()

        listToStr = ', '.join([str(elem) for elem in self.flowShopSchedule.sequence])
        self.sigma.setText(listToStr)
        self.c_max.setText(str(self.flowShopSchedule.makespan))
        self.tt.setText(str(self.flowShopSchedule.totalTardiness))
        
        
    def goBack(self):
        widget.setCurrentIndex(widget.currentIndex()-1)
        widget.removeWidget(self)
        
    def goHome(self):
        widget.setCurrentIndex(0)
        
        for i in range(widget.count(), 0, -1):
            widget.removeWidget(widget.widget(i))
            
        
##################################!
    def displayTFR(self):
        self.plots.setCurrentWidget(self.tfrPlot)
        
        self.tfr.setEnabled(False)
        self.gantt.setEnabled(True)
        self.tar.setEnabled(True)
        
        self.gantt.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(111, 81, 237);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")
        
        self.tfr.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(55, 59, 84);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")
        


        self.tar.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(111, 81, 237);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")
        
        if len(self.flowShopSchedule.stopPreparationRate):        
            self.tap.setEnabled(True)
            
            self.tap.setStyleSheet("color: \"#ffffff\";\n"
    "background-color: rgb(111, 81, 237);\n"
    "border: 1px solid rgb(111, 81, 237) ;\n"
    "border-top-left-radius:\'5px\';\n"
    "border-bottom-left-radius:\'5px\';\n"
    "border-top-right-radius:\'5px\';\n"
    "border-bottom-right-radius:\'5px\';\n"
    "font: 63 11pt \"Montserrat SemiBold\";\n"
    "width: \'95px\';\n"
    "height: \'30px\';")              
        
        plt.clf()
        
        try:
            self.tfrTB.removeWidget(self.toolbar)
            self.plotTFR.removeWidget(self.canv)
            
            sip.delete(self.toolbar)
            sip.delete(self.canv)
            self.toolbar = None
            self.canv = None
            
        except Exception as e:
            print(e)
            pass
    
        self.canv = MatplotlibCanvas(self)
        self.toolbar = NavigationToolbar2QT(self.canv, self.centralwidget)
        self.tfrTB.addWidget(self.toolbar)
        self.plotTFR.addWidget(self.canv)
        
        
        machines = []
        tfr = self.flowShopSchedule.operatingRate
        for i in range(len(tfr)):
                machines.append('M' + str(i+1))

        
        self.canv.gnt.bar(machines,tfr, color=colors)
        
        for index, value in enumerate(tfr):
                self.canv.gnt.text(index, value, str(value)[:6], ha= 'center', va='bottom', color = colors[index])
        self.canv.draw()
##################################!

    def displayTAP(self):  
        self.plots.setCurrentWidget(self.tapPlot)
        
        self.tfr.setEnabled(True)
        self.gantt.setEnabled(True)
        self.tar.setEnabled(True)
        
        self.gantt.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(111, 81, 237);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")
        
    
        
        self.tfr.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(111, 81, 237);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")

        self.tar.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(111, 81, 237);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")   
        
        if len(self.flowShopSchedule.stopPreparationRate): 
        
            self.tap.setEnabled(False)
            self.tap.setStyleSheet("color: \"#ffffff\";\n"
    "background-color: rgb(55, 59, 84);\n"
    "border: 1px solid rgb(111, 81, 237) ;\n"
    "border-top-left-radius:\'5px\';\n"
    "border-bottom-left-radius:\'5px\';\n"
    "border-top-right-radius:\'5px\';\n"
    "border-bottom-right-radius:\'5px\';\n"
    "font: 63 11pt \"Montserrat SemiBold\";\n"
    "width: \'95px\';\n"
    "height: \'30px\';")    
        
        plt.clf()
        
        try:
            self.tapTB.removeWidget(self.toolbar)
            self.plotTAP.removeWidget(self.canv)
            
            sip.delete(self.toolbar)
            sip.delete(self.canv)
            self.toolbar = None
            self.canv = None
            
        except Exception as e:
            print(e)
            pass
    
        self.canv = MatplotlibCanvas(self)
        self.toolbar = NavigationToolbar2QT(self.canv, self.centralwidget)
        self.tapTB.addWidget(self.toolbar)
        self.plotTAP.addWidget(self.canv)
        
        
        machines = []
        tap = self.flowShopSchedule.stopPreparationRate
        for i in range(len(tap)):
                machines.append('M' + str(i+1))

        
        self.canv.gnt.bar(machines,tap, color=colors)
        
        for index, value in enumerate(tap):
                self.canv.gnt.text(index, value, str(value)[:6], ha= 'center', va='bottom', color = colors[index])
        self.canv.draw()


##################################!
    def displayTAR(self):
        self.plots.setCurrentWidget(self.tarPlot)
        
        self.tfr.setEnabled(True)
        self.gantt.setEnabled(True)
        self.tar.setEnabled(False)
        
        self.gantt.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(111, 81, 237);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")
        
        self.tar.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(55, 59, 84);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")
        
        if len(self.flowShopSchedule.stopPreparationRate):            
            self.tap.setEnabled(True)
            self.tap.setStyleSheet("color: \"#ffffff\";\n"
    "background-color: rgb(111, 81, 237);\n"
    "border: 1px solid rgb(111, 81, 237) ;\n"
    "border-top-left-radius:\'5px\';\n"
    "border-bottom-left-radius:\'5px\';\n"
    "border-top-right-radius:\'5px\';\n"
    "border-bottom-right-radius:\'5px\';\n"
    "font: 63 11pt \"Montserrat SemiBold\";\n"
    "width: \'95px\';\n"
    "height: \'30px\';")

        self.tfr.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(111, 81, 237);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")
        
        plt.clf()
        
        try:
            self.tarTB.removeWidget(self.toolbar)
            self.plotTAR.removeWidget(self.canv)
            
            sip.delete(self.toolbar)
            sip.delete(self.canv)
            self.toolbar = None
            self.canv = None
            
        except Exception as e:
            print(e)
            pass
    
        self.canv = MatplotlibCanvas(self)
        self.toolbar = NavigationToolbar2QT(self.canv, self.centralwidget)
        self.tarTB.addWidget(self.toolbar)
        self.plotTAR.addWidget(self.canv)
        
        
        machines = []
        tar = self.flowShopSchedule.stopRate

        for i in range(len(tar)):
                machines.append('M' + str(i+1))
        
        self.canv.gnt.bar(machines,tar, color=colors)
        
        for index, value in enumerate(tar):
                self.canv.gnt.text(index, value, str(value)[:6], ha= 'center', va='bottom', color = colors[index])
        self.canv.draw()
##################################!        
      
##################################!      
    def displayGrid(self):
            self.gridValue = not self.gridValue
            self.displayGantt()

                  
##################################!   

##################################!
    def displayGantt(self):
        self.plots.setCurrentWidget(self.ganttPlot)
        
        self.tfr.setEnabled(True)
        self.gantt.setEnabled(False)
        self.tar.setEnabled(True)
        
        self.tfr.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(111, 81, 237);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")
        
        self.gantt.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(55, 59, 84);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")
        
        if len(self.flowShopSchedule.stopPreparationRate): 
        
            self.tap.setEnabled(True)
            
            self.tap.setStyleSheet("color: \"#ffffff\";\n"
    "background-color: rgb(111, 81, 237);\n"
    "border: 1px solid rgb(111, 81, 237) ;\n"
    "border-top-left-radius:\'5px\';\n"
    "border-bottom-left-radius:\'5px\';\n"
    "border-top-right-radius:\'5px\';\n"
    "border-bottom-right-radius:\'5px\';\n"
    "font: 63 11pt \"Montserrat SemiBold\";\n"
    "width: \'95px\';\n"
    "height: \'30px\';")

        self.tar.setStyleSheet("color: \"#ffffff\";\n"
"background-color: rgb(111, 81, 237);\n"
"border: 1px solid rgb(111, 81, 237) ;\n"
"border-top-left-radius:\'5px\';\n"
"border-bottom-left-radius:\'5px\';\n"
"border-top-right-radius:\'5px\';\n"
"border-bottom-right-radius:\'5px\';\n"
"font: 63 11pt \"Montserrat SemiBold\";\n"
"width: \'95px\';\n"
"height: \'30px\';")       
        
        plt.clf()
        
        
        try:
            self.tarTB.removeWidget(self.toolbar)
            self.plotTAR.removeWidget(self.canv)
            
            sip.delete(self.toolbar)
            sip.delete(self.canv)
            self.toolbar = None
            self.canv = None
            
        except Exception as e:
            print(e)
            pass
    
        self.canv = MatplotlibCanvas(self)
        self.toolbar = NavigationToolbar2QT(self.canv, self.centralwidget)
        self.ganttTB.addWidget(self.toolbar)
        self.plotGANTT.addWidget(self.canv)
        
        self.canv.gnt.set_ylim(0, 50)
 
        self.canv.gnt.set_xlim(0, self.flowShopSchedule.makespan + self.flowShopSchedule.makespan//10)
        
        self.canv.gnt.set_xlabel('')
        self.canv.gnt.set_ylabel('Machines')
        
        

        labels = []
        y_ticks = []
        for i in range(len(self.flowShopSchedule.timeMatrix)):
                labels.append('M' + str(i+1))
                y_ticks.append(15 + i * 10)
        labels.reverse()
        
        self.canv.gnt.set_yticks(y_ticks)


        self.canv.gnt.set_yticklabels(labels)
        
        
        colors = ['#FEB800', '#33D6A0', '#6F52ED', '#FE4C61', '#06A6FF', '#FF7A01', '#FF66B3', '#4DD1BA', '#FA656A', '#477CFF']

        for i in range(0, len(self.flowShopSchedule.timeMatrix)):
                data = [(self.flowShopSchedule.timeMatrixStart[i][j], self.flowShopSchedule.timeMatrix[i][j]-self.flowShopSchedule.timeMatrixStart[i][j]) for j in range(len(self.flowShopSchedule.timeMatrix[0]))]
                self.canv.gnt.broken_barh(data, (10 * len(self.flowShopSchedule.timeMatrix) - 10 * i + 2.5, 5), color=colors)
        
        minor_ticks = np.arange(0, self.flowShopSchedule.makespan + 5, 1)

        if self.gridValue:
                self.ax = plt.gca()
                self.ax.set_xticks(minor_ticks, minor=True)
                
                self.ax.grid(True, axis='x',which="major", alpha=0.2, color='#373B54')
                self.ax.grid(True, axis='x',which="minor", alpha=0.1, color='#373B54')
        
        self.canv.draw()
        
##################################!        
        
    
        

app = QApplication(sys.argv)
start = StartScreen()

widget = QStackedWidget()
widget.addWidget(start)
widget.setFixedHeight(600)
widget.setFixedWidth(800)
widget.setWindowIcon(QtGui.QIcon('assets/Sigma.ico'))
widget.setWindowTitle('Sigma - Flow Shop Scheduling')

   
widget.show()

try:
    sys.exit(app.exec_())
except:
    print("Exiting...")